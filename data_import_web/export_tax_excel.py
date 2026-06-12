# -*- coding: utf-8 -*-
"""
Export tax JSON data to Excel matching the standard template format.
Template: templates/91330300MACULY1C85_税务.xlsx
- Row 1: English column headers
- Row 2: Chinese column headers
- Row 3+: Data
"""

import json
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates')
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, '91330300MACULY1C85_税务.xlsx')


def _load_header_mappings():
    """Extract English→Chinese header mappings from the template Excel."""
    if not os.path.exists(TEMPLATE_FILE):
        return _get_fallback_mappings()

    wb = openpyxl.load_workbook(TEMPLATE_FILE, read_only=True)
    mappings = {}
    for name in wb.sheetnames:
        ws = wb[name]
        en_headers = [cell.value for cell in ws[1]]
        cn_headers = [cell.value for cell in ws[2]]
        # Parse JSON key from sheet name: e.g. "4.纳税申报数据（SBXX_LIST）" → "SBXX_LIST"
        # Supports both half-width () and full-width （） parentheses (sometimes mixed)
        match = re.search(r'[\(（]([A-Za-z_]+)[\)）]', name)
        json_key = match.group(1) if match else None
        if json_key:
            mappings[json_key] = {
                'sheet_name': name,
                'en_headers': en_headers,
                'cn_headers': cn_headers,
            }
    wb.close()
    return mappings


def _get_fallback_mappings():
    """Fallback header mappings if template file is missing."""
    return {
        'NSRJCXX': {
            'sheet_name': '1.企业基本信息(NSRJCXX)',
            'en_headers': ['DJZCLXDM', 'SYKJZDMC', 'JYFW', 'ZCD_DHHM', 'SYKJZD', 'NSRZTDM',
                           'ZCZBBZ', 'NSRZTMC', 'YB', 'ZGY', 'KYRQ', 'SSHYMC', 'DJZCLXMC',
                           'NSLXDM', 'ZCDZ', 'GSZCH', 'DS_DM', 'NSRSBH', 'HYMLDM', 'YYDZ',
                           'HYMLMC', 'LSGXDM', 'DS_MC', 'HZDJRQ', 'DBR_DHHM', 'NSLXMC',
                           'SWJG_DM', 'DHHM', 'ZYRS', 'DBR_YDDHHM', 'LSGXMC', 'ZZJGDM',
                           'SS_MC', 'XYDJ', 'SSHYDM', 'QYHGDM', 'SS_DM', 'NSRMC', 'XYPFFS',
                           'ZCZB_BZMC', 'SCJYQX_Z', 'DBRMC', 'DBR_ZJLX_MC', 'ZCZB', 'XYPFSJ',
                           'DBR_ZJLX_DM', 'SWJG_MC', 'DBR_ZJHM'],
            'cn_headers': ['登记注册类型代码', '适用会计制度名称', '经营范围', '注册地电话号码',
                           '适用会计制度代码', '纳税人状态代码', '注册资本币种', '纳税人状态名称',
                           '邮编', '主管员', '开业日期', '所属行业名称', '登记注册类型名称',
                           '纳税类型代码', '注册地址', '工商注册号', '地市代码', '纳税人识别号',
                           '行业门类代码', '营业地址', '行业门类名称', '隶属关系代码', '地市',
                           '核准登记日期', '法定代表人电话号码', '纳税类型名称', '税务机构代码',
                           '电话号码', '从业人数', '法定代表人移动电话号码', '隶属关系名称',
                           '组织机构代码', '省市', '信用等级', '所属行业代码', '企业海关代码',
                           '省份代码', '纳税人名称', '信用评分分数', '注册资本币种名称',
                           '生产经营期止', '法定代表人名称', '法定代表人证件类型名称', '注册资本',
                           '信用评分时间', '法定代表证件类型代码', '税务机构名称', '法定代表人证件号码'],
        },
        'LXRXX_LIST': {
            'sheet_name': '2.企业联系人信息（LXRXX_LIST）',
            'en_headers': ['DBR_DYDZ', 'DBRMC', 'DBR_DHHM', 'DBR_ZJLX_MC', 'DBR_ZJLX_DM',
                           'DBR_YDDHHM', 'BSSF', 'DBR_ZJHM', 'NSRSBH'],
            'cn_headers': ['代办人电子邮件', '代办人名称', '代办人电话号码', '代办人证件类型名称',
                           '代办人证件类型代码', '代办人移动电话号码', '办税身份', '代办人证件号码',
                           '纳税人识别号'],
        },
        'TZFXX_LIST': {
            'sheet_name': '3.企业投资方信息(TZFXX_LIST)',
            'en_headers': ['TZJE', 'TZBL', 'GJDZ', 'ZJZLMC', 'ZJZLDM', 'TZFJJXZMC',
                           'NSRSBH', 'TZFJJXZDM', 'TZFMC', 'ZJHM'],
            'cn_headers': ['投资金额', '投资比例', '国际地址', '证件种类名称', '证件种类代码',
                           '投资方经济性质名称', '纳税人识别号', '投资方经济性质代码', '投资方名称',
                           '证件号码'],
        },
        'SBXX_LIST': {
            'sheet_name': '4.纳税申报数据（SBXX_LIST）',
            'en_headers': ['SSSQZ', 'SBRQ', 'ZSXMMC', 'YJSE', 'SSSQQ', 'QBXSE', 'YSXSSR',
                           'YNSE', 'SBQX', 'JMSE', 'YBTSE', 'ZSXMDM', 'NSRSBH'],
            'cn_headers': ['所属时间止', '申报日期', '征收项目名称', '预缴税额', '所属时间起',
                           '全部销售额', '应税销售收入', '应纳税额', '申报期限', '减免税额',
                           '应补退税额', '征收项目代码', '纳税人识别号'],
        },
        'ZSXX_LIST': {
            'sheet_name': '5.税款征收信息(ZSXX_LIST)',
            'en_headers': ['SSSQ_Z', 'JKFSRQ', 'SKZL_DM', 'ZSXM_MC', 'SSSQ_Q', 'JKQX',
                           'SKZT_DM', 'ZXPM_DM', 'NSRMC', 'SE', 'SKZL_MC', 'XSSR', 'ZSXM_DM',
                           'SL', 'NSRSBH', 'SKZT_MC'],
            'cn_headers': ['所属时间止', '缴款发生日期', '税款种类代码', '征收项目名称', '所属时间起',
                           '缴款期限', '税款状态代码', '征收品目代码', '纳税人名称', '税额',
                           '税款种类名称', '销售收入', '征收项目代码', '税率', '纳税人识别号',
                           '税款状态名称'],
        },
        'ZCFZBXX_LIST': {
            'sheet_name': '6.企业资产负债表(ZCFZBXX_LIST）',
            'en_headers': ['SKSSQQ', 'BBLX', 'XM', 'NCYE', 'MC', 'SKSSQZ', 'QMYE', 'NSRSBH', 'BSRQ'],
            'cn_headers': ['税款所属期起', '报表类型', '项目', '年初余额', '行次', '税款所属期止',
                           '期末余额', '纳税人识别号', '报送日期'],
        },
        'LRBXX_LIST': {
            'sheet_name': '7.企业利润表(LRBXX_LIST）',
            'en_headers': ['SKSSQQ', 'BBLX', 'XM', 'BQJE', 'MC', 'SKSSQZ', 'BYS', 'NSRSBH',
                           'BSRQ', 'SQJE'],
            'cn_headers': ['税款所属期起', '报表类型', '项目', '本期金额', '行次', '税款所属期止',
                           '本月数', '纳税识别号', '报送日期', '上期金额'],
        },
        'QYWFWZXX_LIST': {
            'sheet_name': '8.企业涉税违法违规信息(QYWFWZXX_LIST）',
            'en_headers': ['NSRSBH', 'DJRQ', 'ZYWFWZSS', 'ZYWFWZSDDM', 'ZYWFWZSDMC',
                           'WFWZLXDM', 'WFWZLXMC', 'WFWZZTDM', 'WFWZZTMC', 'CLCFJDRQ',
                           'LARQ', 'XGZT', 'SSSQQ', 'SSSQZ'],
            'cn_headers': ['纳税人识别号', '登记日期', '主要违法违规事实', '主要违法违规手段代码',
                           '主要违法违规手段名称', '违法违规类型代码', '违法违规类型名称',
                           '违法违规状态代码', '违法违规状态名称', '处理处罚决定日期', '立案日期',
                           '修改状态', '所属时期起', '所属时期止'],
        },
        'SWJCXX_LIST': {
            'sheet_name': '9.企业稽查信息(SWJCXX_LIST）',
            'en_headers': ['NSRSBH', 'AYDJRQ', 'AJLYDM', 'AJLYMC', 'WFWZLXDM', 'WFWZLXMC',
                           'JCLXDM', 'JCLXMC', 'JCZTDM', 'JCZTMC', 'AJCLYJDM', 'AJCLYJMC',
                           'AJMC', 'SSSQQ', 'SSSQZ'],
            'cn_headers': ['纳税人识别号', '案源登记日期', '案件来源代码', '案件来源名称',
                           '违法违规类型代码', '违法违规类型名称', '稽查类型代码', '稽查类型名称',
                           '稽查状态代码', '稽查状态名称', '案件处理意见代码', '案件处理意见名称',
                           '案件名称', '所属时期起', '所属时期止'],
        },
        'QYBGDJXX_LIST': {
            'sheet_name': '10.企业变更登记(QYBGDJXX_LIST）',
            'en_headers': ['NSRSBH', 'BGXMMC', 'BGXMDM', 'BGQNR', 'BGHNR', 'BGRQ'],
            'cn_headers': ['纳税人识别号', '变更项目名称', '变更项目代码', '变更前内容',
                           '变更后内容', '变更日期'],
        },
    }


def _copy_header_style(template_ws, output_ws):
    """Copy header styling (row 1-2) from template sheet to output sheet."""
    for row_idx in [1, 2]:
        for col_idx in range(1, template_ws.max_column + 1):
            src = template_ws.cell(row=row_idx, column=col_idx)
            dst = output_ws.cell(row=row_idx, column=col_idx)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.number_format = src.number_format


# Known tax sheet keys (for detecting tax-standard JSON)
TAX_SHEET_KEYS = {
    'NSRJCXX', 'LXRXX_LIST', 'TZFXX_LIST', 'SBXX_LIST', 'ZSXX_LIST',
    'ZCFZBXX_LIST', 'LRBXX_LIST', 'QYWFWZXX_LIST', 'SWJCXX_LIST', 'QYBGDJXX_LIST',
}

TAX_SHEET_ORDER = [
    'NSRJCXX', 'LXRXX_LIST', 'TZFXX_LIST', 'SBXX_LIST', 'ZSXX_LIST',
    'ZCFZBXX_LIST', 'LRBXX_LIST', 'QYWFWZXX_LIST', 'SWJCXX_LIST', 'QYBGDJXX_LIST',
]


def _write_tax_sheets(output_wb, tax_detail, header_mappings, template_wb):
    """Write tax-standard sheets into the workbook."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True, size=11)
    cn_header_font = Font(bold=True, size=10)
    data_font = Font(size=10)

    for json_key in TAX_SHEET_ORDER:
        record = tax_detail.get(json_key)
        mapping = header_mappings.get(json_key)
        if not mapping:
            continue

        sheet_name = mapping['sheet_name']
        en_headers = mapping['en_headers']
        cn_headers = mapping['cn_headers']

        safe_name = sheet_name[:31]
        ws = output_wb.create_sheet(title=safe_name)

        # Row 1: English headers
        for col_idx, header in enumerate(en_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Chinese headers
        for col_idx, header in enumerate(cn_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = cn_header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        if isinstance(record, dict):
            rows = [record]
        elif isinstance(record, list):
            rows = record
        else:
            rows = []

        for row_idx, row_data in enumerate(rows, 3):
            if not isinstance(row_data, dict):
                continue
            for col_idx, key in enumerate(en_headers, 1):
                value = row_data.get(key, '')
                if value is None:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

        # Set column widths
        if template_wb and sheet_name in template_wb.sheetnames:
            tpl_ws = template_wb[sheet_name]
            for col_idx in range(1, len(en_headers) + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                tpl_col = tpl_ws.column_dimensions.get(col_letter)
                if tpl_col and tpl_col.width:
                    ws.column_dimensions[col_letter].width = tpl_col.width
                else:
                    ws.column_dimensions[col_letter].width = 15
        else:
            for col_idx in range(1, len(en_headers) + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15

        ws.freeze_panes = 'A3'


def _write_generic_sheet(output_wb, data, sheet_name='data'):
    """Fallback: write arbitrary JSON data as a flat sheet."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True, size=11)
    data_font = Font(size=10)

    # Flatten nested dicts
    if isinstance(data, dict):
        rows = []
        for key, value in data.items():
            if isinstance(value, list):
                continue  # skip arrays for generic view
            elif isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    rows.append({'key': f'{key}.{sub_key}', 'value': str(sub_value)})
            else:
                rows.append({'key': key, 'value': str(value)})
    elif isinstance(data, list) and data and isinstance(data[0], dict):
        rows = data
    else:
        rows = [{'data': str(data)}]

    if not rows:
        rows = [{'message': 'No data'}]

    headers = list(rows[0].keys())
    safe_name = sheet_name[:31]
    ws = output_wb.create_sheet(title=safe_name)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            value = row_data.get(key, '')
            if value is None:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    ws.freeze_panes = 'A2'


def _is_tax_data(data):
    """Check if data looks like tax-standard JSON format."""
    if isinstance(data, dict):
        keys = set(data.keys())
        if keys & TAX_SHEET_KEYS:
            return True
    return False


def export_json_to_excel(json_path, output_path):
    """
    Convert a tax standard JSON file to Excel matching the template format.

    Args:
        json_path: Path to the input JSON file, or a dict of data
        output_path: Path for the output Excel file
    """
    if isinstance(json_path, dict):
        data = json_path
    else:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

    # Unwrap taxDetail wrapper if present
    tax_detail = data.get('taxDetail', data)

    header_mappings = _load_header_mappings()

    has_template = os.path.exists(TEMPLATE_FILE)
    template_wb = openpyxl.load_workbook(TEMPLATE_FILE) if has_template else None

    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)

    if _is_tax_data(tax_detail):
        _write_tax_sheets(output_wb, tax_detail, header_mappings, template_wb)
    else:
        _write_generic_sheet(output_wb, tax_detail)

    if template_wb:
        template_wb.close()

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    output_wb.save(output_path)
    return output_path
